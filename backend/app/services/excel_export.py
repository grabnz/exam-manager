"""
Excel export service — generates a workbook from a session's pinned
GridTemplate and its generic score_entries.

Layout (faithful to the original French export):
- one sheet per template *group* (e.g. "Prod. écrite et écriture")
- inside a sheet: each section gets its criteria columns, then Bonus and
  S.T. columns when the section has them; last column = group Total
- final sheet "Note Finale": one column per group + the final score
"""
import io

from . import grid

# Color sets keyed by GridSection.color_key — the first three are the exact
# legacy French palette; the rest cycle for new templates.
PALETTES = {
    "blue":   {"tab": "2E75B6", "sub_hdr": "DAEAF5", "crit_hdr": "C6DFEF",
               "data": "EBF3FB", "total_hdr": "9DC3E6", "total_data": "BDD7EE"},
    "green":  {"tab": "375623", "sub_hdr": "D9EAD3", "crit_hdr": "C6EFCE",
               "data": "EBF5EB", "total_hdr": "70AD47", "total_data": "A9D18E"},
    "orange": {"tab": "843C00", "sub_hdr": "FDE9D9", "crit_hdr": "FCE4D6",
               "data": "FEF0E7", "total_hdr": "ED7D31", "total_data": "F4B183"},
    "purple": {"tab": "5B2D8E", "sub_hdr": "E6DEF2", "crit_hdr": "D9CCEC",
               "data": "F2EDF9", "total_hdr": "9A7FC7", "total_data": "B8A5D6"},
    "teal":   {"tab": "1F6E6B", "sub_hdr": "D6ECEB", "crit_hdr": "C2E3E1",
               "data": "EAF5F4", "total_hdr": "4FA3A0", "total_data": "8CC5C3"},
    "rose":   {"tab": "9C3353", "sub_hdr": "F6DDE5", "crit_hdr": "F0CBD8",
               "data": "FAEDF1", "total_hdr": "C9728F", "total_data": "DA9DB1"},
}
PALETTE_CYCLE = list(PALETTES.keys())
BONUS_BG = "E2EFDA"


def _palette(color_key, index: int) -> dict:
    if color_key in PALETTES:
        return PALETTES[color_key]
    return PALETTES[PALETTE_CYCLE[index % len(PALETTE_CYCLE)]]


def _groups(template):
    """Ordered [(group_key, group_label, [sections])]."""
    out = []
    seen = {}
    for s in template.sections:
        if s.group_key not in seen:
            seen[s.group_key] = (s.group_key, s.group_label, [])
            out.append(seen[s.group_key])
        seen[s.group_key][2].append(s)
    return out


def export_session(session) -> bytes:
    """Takes an ExamSession ORM object (template + entries + students loaded).
    Returns Excel bytes."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    template = session.template
    entries = {e.student_id: e for e in session.entries}
    students = session.class_.students  # ordered by order_index

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(style="thin", color="BBBBBB")

    def mkfill(h): return PatternFill(fill_type="solid", fgColor=h)
    def mkborder(): return Border(left=thin, right=thin, top=thin, bottom=thin)
    def mkalign(h="center"): return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def values_of(student_id) -> dict:
        e = entries.get(student_id)
        return e.values if e else {}

    groups = _groups(template)

    # ── One sheet per group ─────────────────────────────────────────────────
    for gi, (gkey, glabel, sections) in enumerate(groups):
        pal = _palette(sections[0].color_key, gi)
        ws = wb.create_sheet(title=glabel[:31])
        ws.sheet_properties.tabColor = pal["tab"]

        # Column spec: ("crit", section, criterion) / ("bonus"/"subtotal", section, None) / ("total",)
        col_specs = []
        for sec in sections:
            for c in sec.criteria:
                col_specs.append(("crit", sec, c))
            if sec.has_bonus:
                col_specs.append(("bonus", sec, None))
            if sec.has_bonus or sec.allow_st_override:
                col_specs.append(("subtotal", sec, None))
        col_specs.append(("total", None, None))

        NAME_COL  = 1
        DATA_ROW  = 3
        TOTAL_COL = 1 + len(col_specs)

        # section merge spans
        sec_spans = {}
        for ci, (typ, sec, _) in enumerate(col_specs, 2):
            if sec is not None:
                sec_spans.setdefault(sec.id, [sec, ci, ci])
                sec_spans[sec.id][2] = ci

        # Row 1: name (r1-r2), section headers, total (r1-r2)
        ws.merge_cells(start_row=1, start_column=NAME_COL, end_row=2, end_column=NAME_COL)
        c = ws.cell(row=1, column=NAME_COL, value="التلاميذ")
        c.fill, c.font, c.alignment, c.border = (
            mkfill("D9D9D9"), Font(bold=True, size=11), mkalign(), mkborder()
        )
        for sec, s_col, e_col in sec_spans.values():
            if s_col != e_col:
                ws.merge_cells(start_row=1, start_column=s_col, end_row=1, end_column=e_col)
            c = ws.cell(row=1, column=s_col, value=sec.label)
            c.fill = mkfill(pal["sub_hdr"])
            c.font = Font(bold=True, size=10)
            c.alignment, c.border = mkalign(), mkborder()
        ws.merge_cells(start_row=1, start_column=TOTAL_COL, end_row=2, end_column=TOTAL_COL)
        c = ws.cell(row=1, column=TOTAL_COL, value="Total" if template.direction == "ltr" else "المجموع")
        c.fill, c.font, c.alignment, c.border = (
            mkfill(pal["total_hdr"]), Font(bold=True, size=11), mkalign(), mkborder()
        )

        # Row 2: criterion / Bonus / S.T. labels
        for ci, (typ, sec, crit) in enumerate(col_specs, 2):
            if typ == "total":
                break
            lbl = (crit.label if typ == "crit" else
                   "Bonus" if typ == "bonus" else "S.T.")
            clr = (BONUS_BG        if typ == "bonus"    else
                   pal["sub_hdr"]  if typ == "subtotal" else
                   pal["crit_hdr"])
            c = ws.cell(row=2, column=ci, value=lbl)
            c.fill, c.font, c.alignment, c.border = (
                mkfill(clr), Font(bold=True, size=9), mkalign(), mkborder()
            )

        # Student rows
        for ri, student in enumerate(students, DATA_ROW):
            vals = values_of(student.id)
            crit_vals = vals.get("criteria") or {}
            sec_vals  = vals.get("sections") or {}

            c = ws.cell(row=ri, column=NAME_COL, value=student.full_name)
            c.fill, c.font, c.alignment, c.border = (
                mkfill("FFFFFF"), Font(size=9),
                Alignment(horizontal="right", vertical="center"),
                mkborder()
            )
            group_total = 0.0
            for ci, (typ, sec, crit) in enumerate(col_specs, 2):
                cell = ws.cell(row=ri, column=ci)
                if typ == "crit":
                    cell.value = crit_vals.get(crit.id)
                    cell.fill, cell.alignment, cell.border = (
                        mkfill(pal["data"]), mkalign(), mkborder()
                    )
                elif typ == "bonus":
                    cell.value = (sec_vals.get(sec.id) or {}).get("bonus")
                    cell.fill, cell.alignment, cell.border = (
                        mkfill(BONUS_BG), mkalign(), mkborder()
                    )
                elif typ == "subtotal":
                    st_val = grid.section_subtotal(sec, vals)
                    cell.value = st_val or None
                    cell.fill = mkfill(pal["crit_hdr"])
                    cell.font = Font(bold=True, size=9)
                    cell.alignment, cell.border = mkalign(), mkborder()
                else:  # total
                    group_total = sum(grid.section_subtotal(s, vals) for s in sections)
                    cell.value = group_total or None
                    cell.fill = mkfill(pal["total_data"])
                    cell.font = Font(bold=True, size=10)
                    cell.alignment, cell.border = mkalign(), mkborder()

            # sections without bonus/st still contribute via the total column

        # Column widths
        ws.column_dimensions[get_column_letter(NAME_COL)].width = 30
        for ci, (typ, _, __) in enumerate(col_specs, 2):
            ws.column_dimensions[get_column_letter(ci)].width = (
                10 if typ == "total" else 9 if typ in ("subtotal", "bonus") else 6.5
            )
        ws.row_dimensions[1].height = 26
        ws.row_dimensions[2].height = 22
        for ri in range(DATA_ROW, DATA_ROW + len(students)):
            ws.row_dimensions[ri].height = 18
        ws.freeze_panes = ws["B3"]

    # ── Note Finale sheet ────────────────────────────────────────────────────
    final_label = "Moyenne / Note Finale" if template.final_formula == "avg_groups" else "المجموع النهائي"
    ws = wb.create_sheet(title="Note Finale")
    ws.sheet_properties.tabColor = "FFD966"

    DATA_ROW_F  = 2
    NAME_COL_F  = 1
    GROUP_COLS  = list(range(2, 2 + len(groups)))
    FINAL_COL_F = GROUP_COLS[-1] + 1 if GROUP_COLS else 2

    headers_f = [(NAME_COL_F, "التلاميذ", "D9D9D9")]
    for ci, (gi, (gkey, glabel, sections)) in zip(GROUP_COLS, enumerate(groups)):
        pal = _palette(sections[0].color_key, gi)
        headers_f.append((ci, glabel, pal["crit_hdr"]))
    headers_f.append((FINAL_COL_F, final_label, "FFE699"))

    for ci, lbl, clr in headers_f:
        c = ws.cell(row=1, column=ci, value=lbl)
        c.fill, c.font, c.alignment, c.border = (
            mkfill(clr), Font(bold=True, size=10), mkalign(), mkborder()
        )

    for ri, student in enumerate(students, DATA_ROW_F):
        vals = values_of(student.id)
        c = ws.cell(row=ri, column=NAME_COL_F, value=student.full_name)
        c.fill, c.font, c.alignment, c.border = (
            mkfill("FFFFFF"), Font(size=9),
            Alignment(horizontal="right", vertical="center"),
            mkborder()
        )
        gtotals = grid.group_totals(template, vals)
        for ci, (gi, (gkey, glabel, sections)) in zip(GROUP_COLS, enumerate(groups)):
            pal = _palette(sections[0].color_key, gi)
            val = gtotals.get(gkey) or 0.0
            c = ws.cell(row=ri, column=ci, value=val or None)
            c.fill, c.alignment, c.border = (
                mkfill(pal["data"]), mkalign(), mkborder()
            )

        final = grid.final_score(template, vals)
        c = ws.cell(row=ri, column=FINAL_COL_F, value=final)
        c.fill, c.font, c.alignment, c.border = (
            mkfill("FFF2CC"), Font(bold=True, size=11), mkalign(), mkborder()
        )

    ws.column_dimensions["A"].width = 30
    for ci in GROUP_COLS:
        ws.column_dimensions[get_column_letter(ci)].width = 20
    ws.column_dimensions[get_column_letter(FINAL_COL_F)].width = 18
    ws.row_dimensions[1].height = 40
    for ri in range(DATA_ROW_F, DATA_ROW_F + len(students)):
        ws.row_dimensions[ri].height = 18
    ws.freeze_panes = ws["B2"]

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
